"""
Flash Report Mensal — QH Operações
Uso: python flash_mensal.py [YYYY-MM] [--preview] [--enviar]
  YYYY-MM   : mês a analisar (padrão: mês atual)
  --preview : abre HTML no navegador, não envia
  --enviar  : gera PDF e envia por WhatsApp
"""
import sys, os, re, glob, json, subprocess, calendar
from datetime import date, datetime, timedelta
from collections import defaultdict
import psycopg2

# ──────────────────────────────────────────────────────────────────────────────
# CONFIGURAÇÃO
# ──────────────────────────────────────────────────────────────────────────────
sys.stdout.reconfigure(encoding='utf-8')

ENV = r"C:\Users\monit\OneDrive\Área de Trabalho\Ferramenta QH\.env"
_env_raw = open(ENV, encoding='utf-8', errors='ignore').read()
def _env(key):
    if f"{key}=" not in _env_raw:
        return None
    return _env_raw.split(f"{key}=")[1].split()[0]

TERMINAIS = {
    "Manoel Feio":        ["03TR","05TR","07TR","09TR","11TR","20TR","02TR"],
    "GCM — Ítalo Adami":  ["04TR","06TR","15TR","16TR","19VP","34TR"],
    "Estação Itáqua":     ["01TR","21TR","29TR"],
    "Santa Tereza":       ["08TR","10TR","19TR"],
}
EX    = "'97TR','98TR','99TR','99'"
ATIV  = "('Viagem Normal','Viagem Extra')"
TAD, TAI = 8, -5
META_CP, META_PT = 98.0, 92.0
DIFF  = ("CASE WHEN iniciorealizado='' THEN NULL ELSE "
         "EXTRACT(EPOCH FROM (iniciorealizado::timestamp"
         " - inicioprogramado::timestamp))/60 END")
DIFF_ATD = f"FLOOR(({DIFF})) > {TAD}"
DIFF_ADI = f"({DIFF}) < {TAI}"

SAIDA_DIR = r"C:\Users\monit\OneDrive\Área de Trabalho\Ferramenta QH\saidas\mensal"
os.makedirs(SAIDA_DIR, exist_ok=True)

# Normalização de motivos
GRUPOS_MOTIVO = {
    "Trânsito":           ["Operação Atrasada(Transito Congestionado)","Operação Atrasada (Transito Congestionado)","Ope. Atrasada(Necessidade Operacional)"],
    "Falha Mecânica":     ["Falha mecânica.","Falha Mecânica (SOS)","Ope. atrasada por falha mecânica","Ope. adiantada por falha mecânica"],
    "Baixa Estatística":  ["Baixa Estatística","Baixa Estatistica","Baixa estatística"],
    "Falta de Operador":  ["Falta de operador.","Falta de Operador","Adiantado falta de Operador.","Adiantado por falta de operador.","Adiantado por falta de operador","Adiantado falta de Operador.","Adiantado falta de carro"],
    "Falta de Carro":     ["Falta de Carro","Adiantado falta de carro"],
    "Má-Fé":              ["Operação Atrasada(Má fé)","Operação Atrasada (Má Fé)","Atraso Má fé","Ope. Adiantada(Ma Fé)","Ope. Adiantada (Má Fé)","Operação Adiantada (Má Fé)"],
    "Obstrução de Via":   ["Obstrução de Via","Obstrução de via."],
    "Articulação":        ["Articulação Operacional"],
    "Operador Passou Mal":["Operador Passou Mal"],
    "Acidente":           ["Acidente de Trânsito"],
    "Assalto/Violência":  ["Assalto","Ato de Violência","Vandalismo"],
    "Falha Comunicação":  ["Falha de comunicação.","Falha de Comunicação (GPS)","Falha de Comunicação(sombra ponto final)"],
    "Escala Errada":      ["Escala Errada"],
    "Consolidação Manual":["Consolidação Manual"],
}

def normalizar_motivo(m):
    if not m or not m.strip():
        return "Sem registro"
    for grupo, variantes in GRUPOS_MOTIVO.items():
        if any(v.lower() == m.strip().lower() for v in variantes):
            return grupo
    return m.strip()

def linha_terminal(linha):
    for t, linhas in TERMINAIS.items():
        if linha in linhas:
            return t
    return "Outro"

# ──────────────────────────────────────────────────────────────────────────────
# PARSE ARGS
# ──────────────────────────────────────────────────────────────────────────────
PREVIEW = "--preview" in sys.argv
ENVIAR  = "--enviar"  in sys.argv

arg_mes = next((a for a in sys.argv[1:] if re.match(r'\d{4}-\d{2}', a)), None)
if arg_mes:
    ANO, MES = int(arg_mes[:4]), int(arg_mes[5:7])
else:
    hoje = date.today()
    ANO, MES = hoje.year, hoje.month

INICIO = date(ANO, MES, 1)
FIM_MES = date(ANO, MES, calendar.monthrange(ANO, MES)[1])
MES_NOME = ["","Janeiro","Fevereiro","Março","Abril","Maio","Junho",
            "Julho","Agosto","Setembro","Outubro","Novembro","Dezembro"][MES]

# ──────────────────────────────────────────────────────────────────────────────
# BANCO
# ──────────────────────────────────────────────────────────────────────────────
conn = psycopg2.connect(_env("DATABASE_URL"))
cur  = conn.cursor()

# Último dia com dados completos no mês
cur.execute(f"""
SELECT data FROM viagens_qh
WHERE data::date BETWEEN '{INICIO}' AND '{FIM_MES}'
  AND atividade IN {ATIV} AND linha NOT IN ({EX})
  AND iniciorealizado<>'' AND inicioprogramado<>''
GROUP BY data
HAVING COUNT(*) > 100
ORDER BY data DESC LIMIT 1
""")
row = cur.fetchone()
FIM = date.fromisoformat(str(row[0])[:10]) if row else INICIO
TOTAL_DIAS = (FIM - INICIO).days + 1

print(f"Período: {INICIO} a {FIM} ({TOTAL_DIAS} dias)")

# ──────────────────────────────────────────────────────────────────────────────
# QUERIES PRINCIPAIS
# ──────────────────────────────────────────────────────────────────────────────

# 1. CP/PT Geral
cur.execute(f"""
SELECT COUNT(*) as v,
  COUNT(*) FILTER(WHERE iniciorealizado='') as perd,
  COUNT(*) FILTER(WHERE iniciorealizado<>'' AND {DIFF_ATD}) as atd,
  COUNT(*) FILTER(WHERE iniciorealizado<>'' AND {DIFF_ADI}) as adi
FROM viagens_qh
WHERE data::date BETWEEN '{INICIO}' AND '{FIM}'
  AND atividade IN {ATIV} AND linha NOT IN ({EX}) AND inicioprogramado<>''
""")
v,p,atd,adi = (int(x or 0) for x in cur.fetchone())
real = v - p
G = dict(v=v, p=p, atd=atd, adi=adi, real=real,
         cp=round(100*(v-p)/v,1) if v else 0,
         pt=round(100*(real-atd-adi)/real,1) if real else 0)

# 2. Evolução diária
cur.execute(f"""
SELECT data,
  COUNT(*) as v,
  COUNT(*) FILTER(WHERE iniciorealizado='') as p,
  COUNT(*) FILTER(WHERE iniciorealizado<>'' AND {DIFF_ATD}) as atd,
  COUNT(*) FILTER(WHERE iniciorealizado<>'' AND {DIFF_ADI}) as adi
FROM viagens_qh
WHERE data::date BETWEEN '{INICIO}' AND '{FIM}'
  AND atividade IN {ATIV} AND linha NOT IN ({EX}) AND inicioprogramado<>''
GROUP BY data ORDER BY data
""")
EVOLUCAO = []
for d,v2,p2,a2,ad2 in cur.fetchall():
    v2,p2,a2,ad2 = int(v2),int(p2),int(a2),int(ad2)
    r2 = v2-p2
    d_date = date.fromisoformat(str(d)[:10])
    EVOLUCAO.append(dict(
        data=str(d), dia=d_date.day,
        v=v2, p=p2, atd=a2, adi=ad2, real=r2,
        cp=round(100*(v2-p2)/v2,1) if v2 else 0,
        pt=round(100*(r2-a2-ad2)/r2,1) if r2 else 0
    ))

# 3. CP/PT por terminal
TERM_DATA = {}
for t_nome, t_linhas in TERMINAIS.items():
    ln_list = ",".join(f"'{l}'" for l in t_linhas)
    cur.execute(f"""
    SELECT COUNT(*) as v,
      COUNT(*) FILTER(WHERE iniciorealizado='') as p,
      COUNT(*) FILTER(WHERE iniciorealizado<>'' AND {DIFF_ATD}) as atd,
      COUNT(*) FILTER(WHERE iniciorealizado<>'' AND {DIFF_ADI}) as adi
    FROM viagens_qh
    WHERE data::date BETWEEN '{INICIO}' AND '{FIM}'
      AND atividade IN {ATIV} AND linha IN ({ln_list}) AND inicioprogramado<>''
    """)
    v2,p2,a2,ad2 = (int(x or 0) for x in cur.fetchone())
    r2 = v2-p2
    TERM_DATA[t_nome] = dict(
        v=v2, p=p2, atd=a2, adi=ad2, real=r2,
        cp=round(100*(v2-p2)/v2,1) if v2 else 0,
        pt=round(100*(r2-a2-ad2)/r2,1) if r2 else 0,
        linhas=t_linhas
    )

# 4. CP/PT por linha
cur.execute(f"""
SELECT linha,
  COUNT(*) as v,
  COUNT(*) FILTER(WHERE iniciorealizado='') as p,
  COUNT(*) FILTER(WHERE iniciorealizado<>'' AND {DIFF_ATD}) as atd,
  COUNT(*) FILTER(WHERE iniciorealizado<>'' AND {DIFF_ADI}) as adi
FROM viagens_qh
WHERE data::date BETWEEN '{INICIO}' AND '{FIM}'
  AND atividade IN {ATIV} AND linha NOT IN ({EX}) AND inicioprogramado<>''
GROUP BY linha ORDER BY linha
""")
LINHAS = {}
for ln,v2,p2,a2,ad2 in cur.fetchall():
    v2,p2,a2,ad2 = int(v2),int(p2),int(a2),int(ad2)
    r2 = v2-p2
    LINHAS[ln] = dict(
        v=v2, p=p2, atd=a2, adi=ad2, real=r2,
        cp=round(100*(v2-p2)/v2,1) if v2 else 0,
        pt=round(100*(r2-a2-ad2)/r2,1) if r2 else 0,
        terminal=linha_terminal(ln)
    )

# 5. Motivos CP
cur.execute(f"""
SELECT motivo, COUNT(*) as n, COUNT(DISTINCT data) as dias,
  STRING_AGG(DISTINCT linha, ',' ORDER BY linha) as linhas
FROM cco_eventos_cp
WHERE data::date BETWEEN '{INICIO}' AND '{FIM}'
GROUP BY motivo ORDER BY 2 DESC
""")
MOTIVOS_CP_RAW = [dict(motivo=r[0] or '', n=int(r[1]), dias=int(r[2]),
                       linhas=(r[3] or '').split(',')[:5]) for r in cur.fetchall()]

# Agrega por grupo
mc_grupos = defaultdict(lambda: dict(n=0, dias_set=set(), linhas=set()))
for m in MOTIVOS_CP_RAW:
    g = normalizar_motivo(m['motivo'])
    mc_grupos[g]['n'] += m['n']
    mc_grupos[g]['dias_set'].add(m['dias'])
    mc_grupos[g]['linhas'].update(m['linhas'])
MOTIVOS_CP = sorted([dict(grupo=g, n=v['n'], dias=max(v['dias_set']),
                          linhas=sorted(v['linhas'])[:5])
                     for g,v in mc_grupos.items() if g != 'Sem registro'],
                    key=lambda x: -x['n'])

# 6. Motivos PT
cur.execute(f"""
SELECT motivo, COUNT(*) as n, COUNT(DISTINCT date) as dias,
  STRING_AGG(DISTINCT linha, ',' ORDER BY linha) as linhas
FROM cco_eventos_pt
WHERE date::date BETWEEN '{INICIO}' AND '{FIM}'
GROUP BY motivo ORDER BY 2 DESC
""")
mp_grupos = defaultdict(lambda: dict(n=0, dias_set=set(), linhas=set()))
for row in cur.fetchall():
    g = normalizar_motivo(row[0] or '')
    mp_grupos[g]['n'] += int(row[1])
    mp_grupos[g]['dias_set'].add(int(row[2]))
    mp_grupos[g]['linhas'].update((row[3] or '').split(',')[:5])
MOTIVOS_PT = sorted([dict(grupo=g, n=v['n'], dias=max(v['dias_set']),
                          linhas=sorted(v['linhas'])[:5])
                     for g,v in mp_grupos.items() if g != 'Sem registro'],
                    key=lambda x: -x['n'])

# 7. Ofensores CP com recorrência
cur.execute(f"""
SELECT e.motorista, e.linha,
  COUNT(*) as oc, COUNT(DISTINCT e.data) as dias,
  MODE() WITHIN GROUP (ORDER BY e.motivo) as motivo_princ,
  MODE() WITHIN GROUP (ORDER BY EXTRACT(HOUR FROM e.inicioprogramado::timestamp)::int) as hora_princ
FROM cco_eventos_cp e
WHERE e.data::date BETWEEN '{INICIO}' AND '{FIM}'
  AND e.motorista IS NOT NULL AND e.motorista <> ''
  AND e.linha NOT IN ({EX})
GROUP BY e.motorista, e.linha
HAVING COUNT(*) >= 2
ORDER BY dias DESC, oc DESC LIMIT 20
""")
OFENSORES_CP = [dict(mat=r[0], linha=r[1], oc=int(r[2]), dias=int(r[3]),
                     motivo=normalizar_motivo(r[4] or ''), hora=int(r[5] or 0))
                for r in cur.fetchall()]

# 8. Ofensores PT com recorrência
cur.execute(f"""
SELECT e.motorista, e.linha,
  COUNT(*) as oc, COUNT(DISTINCT e.date) as dias,
  MODE() WITHIN GROUP (ORDER BY e.motivo) as motivo_princ,
  MODE() WITHIN GROUP (ORDER BY EXTRACT(HOUR FROM e.inicioprogramado::timestamp)::int) as hora_princ
FROM cco_eventos_pt e
WHERE e.date::date BETWEEN '{INICIO}' AND '{FIM}'
  AND e.motorista IS NOT NULL AND e.motorista <> ''
  AND e.linha NOT IN ({EX})
GROUP BY e.motorista, e.linha
HAVING COUNT(*) >= 3
ORDER BY dias DESC, oc DESC LIMIT 20
""")
OFENSORES_PT = [dict(mat=r[0], linha=r[1], oc=int(r[2]), dias=int(r[3]),
                     motivo=normalizar_motivo(r[4] or ''), hora=int(r[5] or 0))
                for r in cur.fetchall()]

# 9. Análise horária PT
cur.execute(f"""
SELECT EXTRACT(HOUR FROM inicioprogramado::timestamp)::int as hora,
  COUNT(*) as v,
  COUNT(*) FILTER(WHERE iniciorealizado<>'' AND {DIFF_ATD}) as atd,
  COUNT(*) FILTER(WHERE iniciorealizado<>'' AND {DIFF_ADI}) as adi,
  COUNT(*) FILTER(WHERE iniciorealizado='') as perd
FROM viagens_qh
WHERE data::date BETWEEN '{INICIO}' AND '{FIM}'
  AND atividade IN {ATIV} AND linha NOT IN ({EX}) AND inicioprogramado<>''
GROUP BY hora ORDER BY hora
""")
HORARIOS = [dict(h=int(r[0]), v=int(r[1]), atd=int(r[2]),
                 adi=int(r[3]), perd=int(r[4]),
                 irr=int(r[2])+int(r[3])) for r in cur.fetchall()]

# 10. Análise horária por terminal (PT)
HORARIOS_TERM = {}
for t_nome, t_linhas in TERMINAIS.items():
    ln_list = ",".join(f"'{l}'" for l in t_linhas)
    cur.execute(f"""
    SELECT EXTRACT(HOUR FROM inicioprogramado::timestamp)::int as hora,
      COUNT(*) FILTER(WHERE {DIFF_ATD}) as atd,
      COUNT(*) FILTER(WHERE {DIFF_ADI}) as adi,
      COUNT(*) as v
    FROM viagens_qh
    WHERE data::date BETWEEN '{INICIO}' AND '{FIM}'
      AND atividade IN {ATIV} AND linha IN ({ln_list})
      AND inicioprogramado<>'' AND iniciorealizado<>''
    GROUP BY hora ORDER BY hora
    """)
    HORARIOS_TERM[t_nome] = [dict(h=int(r[0]), atd=int(r[1]), adi=int(r[2]), v=int(r[3]))
                              for r in cur.fetchall()]

# 11. Saída de Garagem
cur.execute(f"""
SELECT COUNT(*) as tot,
  COUNT(*) FILTER(WHERE iniciorealizado<>'') as real,
  COUNT(*) FILTER(WHERE iniciorealizado='' ) as nreal,
  COUNT(*) FILTER(WHERE iniciorealizado<>''
    AND iniciorealizado::timestamp < inicioprogramado::timestamp - interval '1 minute') as adiant,
  COUNT(*) FILTER(WHERE iniciorealizado<>''
    AND iniciorealizado::timestamp > inicioprogramado::timestamp + interval '1 minute') as atras,
  COUNT(*) FILTER(WHERE iniciorealizado<>''
    AND iniciorealizado::timestamp BETWEEN inicioprogramado::timestamp - interval '1 minute'
      AND inicioprogramado::timestamp + interval '1 minute') as pont
FROM viagens_qh
WHERE data::date BETWEEN '{INICIO}' AND '{FIM}'
  AND atividade='Saída de Garagem' AND inicioprogramado<>''
  AND linha NOT IN ({EX})
""")
sg = cur.fetchone()
SAIDA_GAR = dict(tot=int(sg[0]), real=int(sg[1]), nreal=int(sg[2]),
                 adiant=int(sg[3]), atras=int(sg[4]), pont=int(sg[5]))

# Por linha
cur.execute(f"""
SELECT linha,
  COUNT(*) as tot,
  COUNT(*) FILTER(WHERE iniciorealizado<>''
    AND iniciorealizado::timestamp > inicioprogramado::timestamp + interval '1 minute') as atras
FROM viagens_qh
WHERE data::date BETWEEN '{INICIO}' AND '{FIM}'
  AND atividade='Saída de Garagem' AND inicioprogramado<>''
  AND linha NOT IN ({EX})
GROUP BY linha HAVING COUNT(*) >= 5 ORDER BY 3 DESC LIMIT 10
""")
SAIDA_GAR_LINHAS = [dict(l=r[0], tot=int(r[1]), atras=int(r[2])) for r in cur.fetchall()]

# 12. Recolhimento
cur.execute(f"""
SELECT status_tempo, COUNT(*) as n,
  AVG(NULLIF(diferenca_tempo_min,'')::numeric) as avg_diff
FROM cco_recolhe
WHERE data::date BETWEEN '{INICIO}' AND '{FIM}'
  AND linha NOT IN ({EX})
GROUP BY status_tempo ORDER BY 2 DESC
""")
RECOLHE = [dict(status=r[0], n=int(r[1]), avg=float(r[2] or 0)) for r in cur.fetchall()]
rec_tot = sum(r['n'] for r in RECOLHE)

# Pior linha recolhe
cur.execute(f"""
SELECT linha, COUNT(*) as n,
  COUNT(*) FILTER(WHERE status_tempo='ACIMA_DO_PROGRAMADO') as atras
FROM cco_recolhe
WHERE data::date BETWEEN '{INICIO}' AND '{FIM}'
  AND linha NOT IN ({EX})
GROUP BY linha HAVING COUNT(*) >= 5
ORDER BY atras DESC LIMIT 5
""")
RECOLHE_LINHAS = [dict(l=r[0], n=int(r[1]), atras=int(r[2])) for r in cur.fetchall()]

# 13. Refeição
cur.execute(f"""
SELECT
  COUNT(*) as tot,
  COUNT(*) FILTER(WHERE duracao_intervalo IS NOT NULL AND duracao_intervalo <> ''
    AND duracao_intervalo <> '00:00') as com_int,
  AVG(CASE WHEN duracao_intervalo ~ '^\\d+:\\d+$'
    THEN EXTRACT(HOUR FROM duracao_intervalo::interval)*60 + EXTRACT(MINUTE FROM duracao_intervalo::interval)
    ELSE NULL END) as avg_min
FROM cco_indicadores_motorista
WHERE data_calendario::date BETWEEN '{INICIO}' AND '{FIM}'
  AND duracao_jornada IS NOT NULL AND duracao_jornada <> ''
""")
rf = cur.fetchone()
REFEICAO = dict(tot=int(rf[0] or 0), com_int=int(rf[1] or 0),
                avg_min=float(rf[2] or 0))

# 14. Aderência de Soltura — Excel de rede
_ADER_TMP = r"C:\Users\monit\AppData\Local\Temp\aderencia_mensal.xlsx"
_PS_ADER  = r"C:\Users\monit\AppData\Local\Temp\copia_aderencia_mensal.ps1"
_MES_STR  = f"{MES:02d}"
with open(_PS_ADER, 'w', encoding='utf-8') as _f:
    _f.write(f"""$cco = "\\\\192.168.211.120\\Ocorrencias\\CCO"
$pasta_ader = Get-ChildItem -LiteralPath $cco | Where-Object {{ $_.Name -like "03*" }} | Select-Object -First 1
$pasta_2026 = Get-ChildItem -LiteralPath $pasta_ader.FullName | Where-Object {{ $_.Name -like "{ANO}*" }} | Select-Object -First 1
$pasta_mes  = Get-ChildItem -LiteralPath $pasta_2026.FullName | Where-Object {{ $_.Name -like "{_MES_STR}*" }} | Select-Object -First 1
if ($pasta_mes) {{
    $xlsx = Get-ChildItem -LiteralPath $pasta_mes.FullName | Where-Object {{ $_.Name -like "*.xlsx" }} | Select-Object -First 1
    if ($xlsx) {{ Copy-Item -LiteralPath $xlsx.FullName -Destination "{_ADER_TMP}" -Force; Write-Host "OK:$($xlsx.Name)" }}
    else {{ Write-Host "NOXLSX" }}
}} else {{ Write-Host "NOPASTA" }}
""")
_ps_out = subprocess.run(["powershell", "-ExecutionPolicy", "Bypass", "-File", _PS_ADER],
                         capture_output=True, text=True, timeout=30)

ADERENCIA = None
if os.path.exists(_ADER_TMP) and "OK:" in _ps_out.stdout:
    try:
        import openpyxl
        _wb = openpyxl.load_workbook(_ADER_TMP, data_only=True)
        if "Resultado" in _wb.sheetnames:
            _ws = _wb["Resultado"]
            _ader_data = {}
            for _row in _ws.iter_rows(min_row=2, values_only=True):
                if _row[0] and _row[1] is not None:
                    _ader_data[str(_row[0]).strip()] = _row[1]
            ADERENCIA = _ader_data
            print(f"Aderência: {len(_ader_data)} registros da aba Resultado")
    except Exception as _e:
        print(f"Aderência: erro ao ler Excel — {_e}")
else:
    print(f"Aderência: Excel não disponível ({_ps_out.stdout.strip() or 'sem saída'})")

# ──────────────────────────────────────────────────────────────────────────────
# ANALYTICS — OPORTUNIDADES
# ──────────────────────────────────────────────────────────────────────────────

def score_opp(oc, dias, total_dias, conc=0.5):
    """Score de oportunidade: impacto + recorrência + concentração"""
    imp  = min(oc / 30.0, 1.0)
    rec  = min(dias / total_dias, 1.0)
    return imp * 0.45 + rec * 0.40 + conc * 0.15

def prioridade(s):
    if s >= 0.55: return "ALTA",    "#DC2626", "🔴"
    if s >= 0.28: return "MÉDIA",   "#D97706", "🟡"
    return                "BAIXA",  "#15803D", "🟢"

# Oportunidades PT por linha
OPPS = []
for ln, ld in LINHAS.items():
    irr = ld['atd'] + ld['adi']
    if irr < 5: continue
    # Concentração horária (estimativa via horarios gerais)
    s = score_opp(irr, TOTAL_DIAS, TOTAL_DIAS)
    pri, cor, ico = prioridade(s)
    OPPS.append(dict(tipo='PT', linha=ln, terminal=ld['terminal'],
                     oc=irr, atd=ld['atd'], adi=ld['adi'],
                     pt=ld['pt'], cp=ld['cp'],
                     score=s, pri=pri, cor=cor, ico=ico))

# Oportunidades CP por linha
for ln, ld in LINHAS.items():
    if ld['p'] < 3: continue
    s = score_opp(ld['p'], TOTAL_DIAS, TOTAL_DIAS)
    pri, cor, ico = prioridade(s)
    OPPS.append(dict(tipo='CP', linha=ln, terminal=ld['terminal'],
                     oc=ld['p'], atd=0, adi=0,
                     pt=ld['pt'], cp=ld['cp'],
                     score=s, pri=pri, cor=cor, ico=ico))

OPPS.sort(key=lambda x: -x['score'])
TOP_OPPS = OPPS[:10]

conn.close()

print(f"CP={G['cp']}% PT={G['pt']}% | Perdidas={G['p']} Atrasos={G['atd']} Adiant={G['adi']}")
print(f"Oportunidades identificadas: {len(OPPS)} (top {len(TOP_OPPS)} listadas)")

# ──────────────────────────────────────────────────────────────────────────────
# HELPERS SVG/HTML
# ──────────────────────────────────────────────────────────────────────────────

def fp(v): return f"{v:.1f}%"
def fn(v): return f"{v:,}".replace(",",".")

def cor_pct(v, meta):
    if v >= meta:   return "#15803D"
    if v >= meta-3: return "#D97706"
    return "#DC2626"

def svg_bars_h(items, w=420, bar_h=22, pad=4, cor="#1BBEAA", max_val=None):
    """Gráfico de barras horizontais: items = [(label, val, cor_opcional)]"""
    if not items: return ""
    max_v = max_val or max(x[1] for x in items) or 1
    row_h = bar_h + pad
    h = row_h * len(items) + 10
    lines = [f'<svg viewBox="0 0 {w} {h}" xmlns="http://www.w3.org/2000/svg" style="width:100%;max-width:{w}px">']
    for i, item in enumerate(items):
        lbl, val = item[0], item[1]
        c = item[2] if len(item) > 2 else cor
        y = i * row_h + 2
        bw = int((val / max_v) * (w - 160))
        lines.append(f'<text x="0" y="{y+bar_h-5}" font-size="10" fill="#374151" font-family="Segoe UI,Arial">{lbl[:22]}</text>')
        lines.append(f'<rect x="120" y="{y}" width="{bw}" height="{bar_h-pad}" rx="3" fill="{c}" opacity="0.85"/>')
        lines.append(f'<text x="{120+bw+4}" y="{y+bar_h-5}" font-size="10" fill="#374151" font-family="Segoe UI,Arial" font-weight="600">{val}</text>')
    lines.append("</svg>")
    return "\n".join(lines)

def svg_line(data_list, label_cp="CP%", label_pt="PT%", meta_cp=META_CP, meta_pt=META_PT, w=560, h=160):
    """Gráfico de linha duplo CP e PT ao longo do mês."""
    if not data_list: return ""
    n = len(data_list)
    pad_l, pad_r, pad_t, pad_b = 40, 10, 10, 30
    cw = w - pad_l - pad_r
    ch = h - pad_t - pad_b
    mn, mx = 70, 100
    def sx(i): return pad_l + i * cw / max(n-1, 1)
    def sy(v): return pad_t + ch - (v - mn) / (mx - mn) * ch

    lines = [f'<svg viewBox="0 0 {w} {h}" xmlns="http://www.w3.org/2000/svg" style="width:100%;max-width:{w}px">']
    # Grid
    for v in range(int(mn), int(mx)+1, 5):
        y = sy(v)
        lines.append(f'<line x1="{pad_l}" y1="{y:.1f}" x2="{w-pad_r}" y2="{y:.1f}" stroke="#E2E8F0" stroke-width="1"/>')
        lines.append(f'<text x="{pad_l-4}" y="{y+3:.1f}" font-size="8" fill="#9CA3AF" text-anchor="end" font-family="Segoe UI,Arial">{v}</text>')
    # Meta lines
    for meta, cor_m, lbl_m in [(meta_cp,"#15803D",label_cp),(meta_pt,"#1D4ED8",label_pt)]:
        ym = sy(meta)
        lines.append(f'<line x1="{pad_l}" y1="{ym:.1f}" x2="{w-pad_r}" y2="{ym:.1f}" stroke="{cor_m}" stroke-width="1" stroke-dasharray="4,3" opacity="0.4"/>')
    # CP line
    pts_cp = " ".join(f"{sx(i):.1f},{sy(d['cp']):.1f}" for i,d in enumerate(data_list))
    lines.append(f'<polyline points="{pts_cp}" stroke="#15803D" stroke-width="2" fill="none" stroke-linejoin="round"/>')
    # PT line
    pts_pt = " ".join(f"{sx(i):.1f},{sy(d['pt']):.1f}" for i,d in enumerate(data_list))
    lines.append(f'<polyline points="{pts_pt}" stroke="#1D4ED8" stroke-width="2" fill="none" stroke-linejoin="round"/>')
    # Labels dias (cada 5)
    for i, d in enumerate(data_list):
        if i % 5 == 0 or i == n-1:
            lines.append(f'<text x="{sx(i):.1f}" y="{h-5}" font-size="8" fill="#9CA3AF" text-anchor="middle" font-family="Segoe UI,Arial">{d["dia"]}</text>')
    # Legenda
    lines.append(f'<rect x="{pad_l}" y="2" width="10" height="6" fill="#15803D" rx="1"/>')
    lines.append(f'<text x="{pad_l+13}" y="9" font-size="8" fill="#15803D" font-family="Segoe UI,Arial" font-weight="600">CP%</text>')
    lines.append(f'<rect x="{pad_l+45}" y="2" width="10" height="6" fill="#1D4ED8" rx="1"/>')
    lines.append(f'<text x="{pad_l+58}" y="9" font-size="8" fill="#1D4ED8" font-family="Segoe UI,Arial" font-weight="600">PT%</text>')
    lines.append("</svg>")
    return "\n".join(lines)

def svg_heat_horas(horarios, w=500, h=60):
    """Mini heatmap de irregularidades por hora."""
    if not horarios: return ""
    max_irr = max(h2['irr'] for h2 in horarios) or 1
    cw2 = w / 24
    lines = [f'<svg viewBox="0 0 {w} {h}" xmlns="http://www.w3.org/2000/svg" style="width:100%;max-width:{w}px">']
    for hr in range(24):
        hd = next((h2 for h2 in horarios if h2['h'] == hr), None)
        irr = hd['irr'] if hd else 0
        ratio = irr / max_irr
        r = int(220 + ratio * 35)
        g = int(220 - ratio * 160)
        b = int(220 - ratio * 200)
        x = hr * cw2
        lines.append(f'<rect x="{x:.1f}" y="0" width="{cw2:.1f}" height="35" fill="rgb({r},{g},{b})" rx="1"/>')
        lines.append(f'<text x="{x+cw2/2:.1f}" y="47" font-size="7" fill="#6B7280" text-anchor="middle" font-family="Segoe UI,Arial">{hr:02d}h</text>')
        if irr > 0:
            lines.append(f'<text x="{x+cw2/2:.1f}" y="22" font-size="7" fill="#374151" text-anchor="middle" font-family="Segoe UI,Arial">{irr}</text>')
    lines.append("</svg>")
    return "\n".join(lines)

# ──────────────────────────────────────────────────────────────────────────────
# CSS
# ──────────────────────────────────────────────────────────────────────────────
CSS = """
*,*::before,*::after{box-sizing:border-box;margin:0;padding:0}
body{font-family:'Segoe UI',Arial,sans-serif;background:#fff;color:#111827;font-size:12px;line-height:1.5;-webkit-print-color-adjust:exact;print-color-adjust:exact}
.page{width:210mm;min-height:297mm;margin:0 auto;padding:12mm 14mm 10mm;page-break-after:always;position:relative}
.page:last-child{page-break-after:avoid}
@media print{.page{box-shadow:none;margin:0}}
@media screen{.page{box-shadow:0 2px 8px rgba(0,0,0,.12);margin:0 auto 20px}}

/* HEADER */
.hdr{display:flex;justify-content:space-between;align-items:flex-start;border-bottom:3px solid #1BBEAA;padding-bottom:8px;margin-bottom:14px}
.hdr-l{flex:1}
.hdr-title{font-size:18px;font-weight:800;color:#1A3252;letter-spacing:-.3px}
.hdr-title span{color:#1BBEAA}
.hdr-sub{font-size:10px;color:#6B7280;margin-top:2px}
.date-box{background:#1A3252;color:#fff;border-radius:6px;padding:6px 12px;text-align:center;min-width:80px}
.date-box-l{font-size:7px;text-transform:uppercase;letter-spacing:.1em;color:#93C5FD;margin-bottom:1px}
.date-box-v{font-size:13px;font-weight:800;white-space:nowrap}

/* COVER */
.cover{display:flex;flex-direction:column;justify-content:center;align-items:center;min-height:240mm;background:linear-gradient(155deg,#1A3252 0%,#0F2644 100%);border-radius:8px;color:#fff;text-align:center;padding:30px}
.cover-icon{font-size:48px;margin-bottom:16px}
.cover-title{font-size:28px;font-weight:800;letter-spacing:-.5px;line-height:1.2;margin-bottom:8px}
.cover-title span{color:#1BBEAA}
.cover-sub{font-size:14px;color:rgba(255,255,255,.6);margin-bottom:24px}
.cover-period{background:rgba(255,255,255,.1);border:1px solid rgba(255,255,255,.2);border-radius:8px;padding:12px 24px;font-size:13px;color:#fff;margin-bottom:8px}
.cover-status{font-size:11px;color:#4ADE80;font-weight:600}

/* SECTION */
.sec{font-size:10px;font-weight:700;text-transform:uppercase;letter-spacing:.08em;color:#6B7280;margin:12px 0 8px;display:flex;align-items:center;gap:8px}
.sec::after{content:'';flex:1;height:1px;background:#E2E8F0}
.sec.ok::before{content:'';width:10px;height:10px;border-radius:2px;background:#1BBEAA;display:inline-block}

/* KPI */
.kpi-row{display:flex;gap:8px;margin-bottom:10px;flex-wrap:wrap}
.kpi{flex:1;min-width:80px;background:#F8FAFC;border:1px solid #E2E8F0;border-radius:7px;padding:9px 11px}
.kpi.accent{border-left:4px solid}
.kpi-l{font-size:8px;text-transform:uppercase;letter-spacing:.06em;color:#6B7280;margin-bottom:2px}
.kpi-v{font-size:28px;font-weight:800;line-height:1}
.kpi-v.sm{font-size:20px}
.kpi-s{font-size:9px;color:#6B7280;margin-top:2px}

/* TABLE */
table{width:100%;border-collapse:collapse;font-size:11px;margin-bottom:10px}
th{background:#1E3A5F;color:#fff;padding:5px 8px;text-align:left;font-size:9px;text-transform:uppercase;letter-spacing:.04em;font-weight:700}
th.r,td.r{text-align:right;font-variant-numeric:tabular-nums}
td{padding:4px 8px;border-bottom:1px solid #F1F5F9;vertical-align:middle}
tr:nth-child(even) td{background:#FAFBFC}
tr:last-child td{border-bottom:none}
.bd{display:inline-block;padding:1px 7px;border-radius:8px;font-size:10px;font-weight:700}
.bd-ok{background:#DCFCE7;color:#15803D}
.bd-w{background:#FEF3C7;color:#B45309}
.bd-c{background:#FEE2E2;color:#DC2626}
.bd-b{background:#DBEAFE;color:#1D4ED8}

/* OPP CARDS */
.opp-list{display:flex;flex-direction:column;gap:7px;margin-bottom:10px}
.opp{border-radius:7px;padding:9px 12px;border-left:4px solid;background:#FAFBFC;border:1px solid #E2E8F0}
.opp-hdr{display:flex;align-items:center;gap:8px;margin-bottom:4px;flex-wrap:wrap}
.opp-n{width:20px;height:20px;border-radius:50%;background:#1A3252;color:#fff;font-size:10px;font-weight:800;display:flex;align-items:center;justify-content:center;flex-shrink:0}
.opp-title{font-weight:700;font-size:12px;flex:1}
.opp-pri{font-size:10px;font-weight:700;padding:2px 8px;border-radius:8px}
.opp-det{font-size:10px;color:#6B7280;display:flex;gap:12px;flex-wrap:wrap}
.opp-det span{display:flex;gap:3px;align-items:center}
.opp-det strong{color:#374151}

/* PRIORITY BOX */
.atacar{border-radius:7px;padding:9px 12px;margin-bottom:6px;border-left:4px solid}
.atacar.a{background:#FFF1F2;border-color:#DC2626}
.atacar.m{background:#FFFBEB;border-color:#D97706}
.atacar.b{background:#F0FDF4;border-color:#15803D}
.atacar-hdr{display:flex;align-items:center;gap:6px;font-weight:700;font-size:11px;margin-bottom:3px}
.atacar-det{font-size:10px;color:#6B7280;display:flex;gap:10px;flex-wrap:wrap}

/* FOOTER */
.ftr{position:absolute;bottom:8mm;left:14mm;right:14mm;border-top:1px solid #E2E8F0;padding-top:4px;font-size:8px;color:#9CA3AF;display:flex;justify-content:space-between}

/* TERMINAL CARD */
.tcard{border:1px solid #E2E8F0;border-radius:8px;margin-bottom:10px;overflow:hidden}
.tcard-hdr{background:#1A3252;color:#fff;padding:8px 12px;display:flex;align-items:center;gap:10px;flex-wrap:wrap}
.tcard-nome{font-size:13px;font-weight:800;flex:1;text-transform:uppercase;letter-spacing:.03em}
.tkpi{text-align:center;background:rgba(255,255,255,.12);border-radius:5px;padding:4px 10px;min-width:60px}
.tkpi-l{font-size:7px;color:rgba(255,255,255,.5);text-transform:uppercase;letter-spacing:.07em}
.tkpi-v{font-size:16px;font-weight:800}
.tcard-body{padding:10px 12px}
"""

# ──────────────────────────────────────────────────────────────────────────────
# HTML SECTIONS
# ──────────────────────────────────────────────────────────────────────────────
MES_ANO_STR = f"{MES_NOME.upper()} / {ANO}"
PERIODO_STR = f"{INICIO.strftime('%d/%m/%Y')} a {FIM.strftime('%d/%m/%Y')}"
GERADO_EM   = datetime.now().strftime("%d/%m/%Y %H:%M")

def hdr(subtitulo=""):
    return f"""<div class="hdr">
      <div class="hdr-l">
        <div class="hdr-title">FLASH REPORT <span>MENSAL</span></div>
        <div class="hdr-sub">{subtitulo or MES_NOME+' '+str(ANO)+' — '+PERIODO_STR}</div>
      </div>
      <div class="date-box"><div class="date-box-l">PERÍODO</div>
        <div class="date-box-v">{MES_NOME[:3].upper()}/{ANO}</div></div>
    </div>"""

def ftr(pg):
    return f"""<div class="ftr">
      <span>QH Operações · Flash Report Mensal · {MES_NOME} {ANO}</span>
      <span>Atualizado em {GERADO_EM} · Pág. {pg}</span>
    </div>"""

# ── Página 1: Capa ────────────────────────────────────────────────────────────
P_CAPA = f"""<div class="page">
  <div class="cover">
    <div class="cover-icon">📊</div>
    <div class="cover-title">FLASH REPORT <span>MENSAL</span></div>
    <div class="cover-sub">QH Operações — Relatório Analítico</div>
    <div class="cover-period">
      <strong>{MES_NOME.upper()} / {ANO}</strong><br>
      Período: {PERIODO_STR}<br>
      {TOTAL_DIAS} dias analisados
    </div>
    <div class="cover-status">✅ DADOS ATUALIZADOS EM {GERADO_EM}</div>
  </div>
  {ftr(1)}
</div>"""

# ── Página 2: Resumo Executivo ────────────────────────────────────────────────
grafico_evolucao = svg_line(EVOLUCAO)
cor_cp = cor_pct(G['cp'], META_CP)
cor_pt = cor_pct(G['pt'], META_PT)

# Mini tabela resumo por terminal
linhas_term = ""
for t, td in TERM_DATA.items():
    c_cp = cor_pct(td['cp'], META_CP)
    c_pt = cor_pct(td['pt'], META_PT)
    linhas_term += f"""<tr>
      <td><strong>{t}</strong></td>
      <td class="r"><span class="bd {'bd-ok' if td['cp']>=META_CP else 'bd-w' if td['cp']>=META_CP-3 else 'bd-c'}">{fp(td['cp'])}</span></td>
      <td class="r"><span class="bd {'bd-ok' if td['pt']>=META_PT else 'bd-w' if td['pt']>=META_PT-3 else 'bd-c'}">{fp(td['pt'])}</span></td>
      <td class="r">{fn(td['v'])}</td><td class="r" style="color:#DC2626;font-weight:{'700' if td['p']>0 else '400'}">{td['p']}</td>
      <td class="r">{td['atd']}</td><td class="r">{td['adi']}</td>
    </tr>"""

P_RESUMO = f"""<div class="page">
  {hdr("Resumo Executivo")}
  <div class="kpi-row">
    <div class="kpi accent" style="border-left-color:{cor_cp};flex:2">
      <div class="kpi-l">CP — Cumprimento de Partida</div>
      <div class="kpi-v" style="color:{cor_cp}">{fp(G['cp'])}</div>
      <div class="kpi-s">Meta {META_CP}% &nbsp;·&nbsp; {MES_NOME} {ANO}</div>
    </div>
    <div class="kpi accent" style="border-left-color:{cor_pt};flex:2">
      <div class="kpi-l">PT — Pontualidade</div>
      <div class="kpi-v" style="color:{cor_pt}">{fp(G['pt'])}</div>
      <div class="kpi-s">Meta {META_PT}% &nbsp;·&nbsp; {MES_NOME} {ANO}</div>
    </div>
  </div>
  <div class="kpi-row">
    <div class="kpi"><div class="kpi-l">Previstas</div><div class="kpi-v sm" style="color:#1E40AF">{fn(G['v'])}</div></div>
    <div class="kpi"><div class="kpi-l">Realizadas</div><div class="kpi-v sm" style="color:#15803D">{fn(G['real'])}</div></div>
    <div class="kpi"><div class="kpi-l">Perdidas</div><div class="kpi-v sm" style="color:#DC2626">{G['p']}</div></div>
    <div class="kpi"><div class="kpi-l">Atrasos &gt;{TAD}min</div><div class="kpi-v sm" style="color:#D97706">{G['atd']}</div></div>
    <div class="kpi"><div class="kpi-l">Adiantamentos &gt;{abs(TAI)}min</div><div class="kpi-v sm" style="color:#7C3AED">{G['adi']}</div></div>
  </div>
  <div class="sec">Evolução Diária — CP e PT</div>
  {grafico_evolucao}
  <div class="sec">Resumo por Terminal</div>
  <table>
    <thead><tr><th>Terminal</th><th class="r">CP%</th><th class="r">PT%</th>
      <th class="r">Previstas</th><th class="r">Perdidas</th>
      <th class="r">Atrasos</th><th class="r">Adiant.</th></tr></thead>
    <tbody>{linhas_term}</tbody>
  </table>
  {ftr(2)}
</div>"""

# ── Página 3: Top Oportunidades + Onde Atacar ─────────────────────────────────
opp_html = ""
for i, o in enumerate(TOP_OPPS[:8]):
    tipo_badge = f'<span class="bd bd-c">PT</span>' if o['tipo']=='PT' else f'<span class="bd bd-b">CP</span>'
    opp_html += f"""<div class="opp" style="border-left-color:{o['cor']}">
      <div class="opp-hdr">
        <div class="opp-n">{i+1}</div>
        <div class="opp-title">{o['linha']} {tipo_badge}</div>
        <span class="opp-pri" style="background:{o['cor']}22;color:{o['cor']}">{o['ico']} {o['pri']}</span>
      </div>
      <div class="opp-det">
        <span>Terminal: <strong>{o['terminal']}</strong></span>
        {'<span>PT: <strong>'+fp(o['pt'])+'</strong></span>' if o['tipo']=='PT' else ''}
        {'<span>Atrasos: <strong>'+str(o['atd'])+'</strong></span>' if o['atd'] else ''}
        {'<span>Adiant.: <strong>'+str(o['adi'])+'</strong></span>' if o['adi'] else ''}
        {'<span>Perdas: <strong>'+str(o['oc'])+'</strong></span>' if o['tipo']=='CP' else ''}
      </div>
    </div>"""

atacar_html = ""
for o in [x for x in TOP_OPPS if x['pri']=='ALTA'][:5]:
    cls = 'a' if o['pri']=='ALTA' else 'm' if o['pri']=='MÉDIA' else 'b'
    atacar_html += f"""<div class="atacar {cls}">
      <div class="atacar-hdr">{o['ico']} {o['pri']} — {o['linha']} ({o['tipo']})</div>
      <div class="atacar-det">
        <span>Terminal: <strong>{o['terminal']}</strong></span>
        <span>Ocorrências: <strong>{o['oc']}</strong></span>
        {'<span>PT atual: <strong>'+fp(o['pt'])+'</strong> (meta '+fp(META_PT)+')</span>' if o['tipo']=='PT' else ''}
      </div>
    </div>"""
if not atacar_html:
    atacar_html = '<p style="color:#15803D;font-size:11px;font-weight:600">✅ Nenhuma linha com prioridade ALTA no período.</p>'

P_OPPS = f"""<div class="page">
  {hdr("Oportunidades e Direcionamento")}
  <div class="sec">Top Oportunidades do Mês — Impacto + Recorrência + Concentração</div>
  <div class="opp-list">{opp_html}</div>
  <div class="sec">Onde Atacar</div>
  {atacar_html}
  {ftr(3)}
</div>"""

# ── Páginas 4-7: Análise por Terminal ────────────────────────────────────────
PGS_TERM = []
pg_num = 4
for t_nome, td in TERM_DATA.items():
    t_linhas_data = [(ln, LINHAS[ln]) for ln in td['linhas'] if ln in LINHAS]
    t_linhas_data.sort(key=lambda x: x[1]['pt'])

    linhas_rows = ""
    for ln, ld in t_linhas_data:
        c_cp = 'bd-ok' if ld['cp']>=META_CP else 'bd-w' if ld['cp']>=META_CP-3 else 'bd-c'
        c_pt = 'bd-ok' if ld['pt']>=META_PT else 'bd-w' if ld['pt']>=META_PT-3 else 'bd-c'
        linhas_rows += f"""<tr>
          <td><strong>{ln}</strong></td>
          <td class="r"><span class="bd {c_cp}">{fp(ld['cp'])}</span></td>
          <td class="r"><span class="bd {c_pt}">{fp(ld['pt'])}</span></td>
          <td class="r">{fn(ld['v'])}</td>
          <td class="r" style="{'color:#DC2626;font-weight:700' if ld['p']>0 else ''}">{ld['p']}</td>
          <td class="r">{ld['atd']}</td><td class="r">{ld['adi']}</td>
        </tr>"""

    # Horários críticos do terminal
    ht = HORARIOS_TERM.get(t_nome, [])
    ht_sorted = sorted(ht, key=lambda x: -(x['atd']+x['adi']))[:3]
    horarios_crit = ", ".join(f"{x['h']:02d}h ({x['atd']+x['adi']} irr.)" for x in ht_sorted) or "—"

    c_cp_t = cor_pct(td['cp'], META_CP)
    c_pt_t = cor_pct(td['pt'], META_PT)

    PGS_TERM.append(f"""<div class="page">
      {hdr(f"Análise por Terminal — {t_nome}")}
      <div class="tcard">
        <div class="tcard-hdr">
          <div class="tcard-nome">{t_nome}</div>
          <div class="tkpi"><div class="tkpi-l">CP</div><div class="tkpi-v" style="color:{'#4ADE80' if td['cp']>=META_CP else '#FCD34D'}">{fp(td['cp'])}</div></div>
          <div class="tkpi"><div class="tkpi-l">PT</div><div class="tkpi-v" style="color:{'#4ADE80' if td['pt']>=META_PT else '#FCD34D'}">{fp(td['pt'])}</div></div>
          <div class="tkpi"><div class="tkpi-l">Previstas</div><div class="tkpi-v" style="color:#93C5FD">{fn(td['v'])}</div></div>
          <div class="tkpi"><div class="tkpi-l">Perdidas</div><div class="tkpi-v" style="color:{'#FCA5A5' if td['p']>0 else '#4ADE80'}">{td['p']}</div></div>
        </div>
        <div class="tcard-body">
          <div class="sec">Linhas do Terminal</div>
          <table><thead><tr><th>Linha</th><th class="r">CP%</th><th class="r">PT%</th>
            <th class="r">Previstas</th><th class="r">Perdidas</th>
            <th class="r">Atrasos</th><th class="r">Adiant.</th></tr></thead>
            <tbody>{linhas_rows}</tbody></table>
          <div class="sec">Horários Críticos (PT)</div>
          <p style="font-size:11px;color:#374151;margin-bottom:8px">{horarios_crit}</p>
        </div>
      </div>
      {ftr(pg_num)}
    </div>""")
    pg_num += 1

# ── Página: Motivos CP e PT ──────────────────────────────────────────────────
mc_items = [(m['grupo'], m['n'], "#DC2626") for m in MOTIVOS_CP[:12]]
mp_items = [(m['grupo'], m['n'], "#D97706") for m in MOTIVOS_PT[:12]]
mc_max = max((x[1] for x in mc_items), default=1)
mp_max = max((x[1] for x in mp_items), default=1)

P_MOTIVOS = f"""<div class="page">
  {hdr("Motivos e Causas")}
  <div class="sec">Motivos de Perda — CP ({sum(m['n'] for m in MOTIVOS_CP)} eventos)</div>
  {svg_bars_h(mc_items, cor="#DC2626", max_val=mc_max)}
  <div class="sec">Motivos de Irregularidade — PT ({sum(m['n'] for m in MOTIVOS_PT)} eventos)</div>
  {svg_bars_h(mp_items, cor="#D97706", max_val=mp_max)}
  {ftr(pg_num)}
</div>"""
pg_num += 1

# ── Página: Ofensores e Recorrência ──────────────────────────────────────────
of_cp_rows = ""
for o in OFENSORES_CP[:10]:
    of_cp_rows += f"""<tr>
      <td><strong>{o['mat']}</strong></td><td>{o['linha']}</td>
      <td class="r">{o['oc']}</td><td class="r"><strong>{o['dias']}</strong></td>
      <td>{o['motivo']}</td><td class="r">{o['hora']:02d}h</td>
    </tr>"""
of_pt_rows = ""
for o in OFENSORES_PT[:10]:
    of_pt_rows += f"""<tr>
      <td><strong>{o['mat']}</strong></td><td>{o['linha']}</td>
      <td class="r">{o['oc']}</td><td class="r"><strong>{o['dias']}</strong></td>
      <td>{o['motivo']}</td><td class="r">{o['hora']:02d}h</td>
    </tr>"""

P_OFENSORES = f"""<div class="page">
  {hdr("Ofensores e Recorrência")}
  <div class="sec">Ofensores CP — Operadores com ≥ 2 Perdas</div>
  <table>
    <thead><tr><th>Matrícula</th><th>Linha</th><th class="r">Ocorr.</th>
      <th class="r">Dias</th><th>Motivo Princ.</th><th class="r">Hora</th></tr></thead>
    <tbody>{of_cp_rows or '<tr><td colspan="6" style="text-align:center;color:#6B7280">Sem ofensores com ≥ 2 perdas</td></tr>'}</tbody>
  </table>
  <div class="sec">Ofensores PT — Operadores com ≥ 3 Irregularidades</div>
  <table>
    <thead><tr><th>Matrícula</th><th>Linha</th><th class="r">Ocorr.</th>
      <th class="r">Dias</th><th>Motivo Princ.</th><th class="r">Hora</th></tr></thead>
    <tbody>{of_pt_rows or '<tr><td colspan="6" style="text-align:center;color:#6B7280">Sem ofensores com ≥ 3 irregularidades</td></tr>'}</tbody>
  </table>
  {ftr(pg_num)}
</div>"""
pg_num += 1

# ── Página: Análise Horária ──────────────────────────────────────────────────
hor_sorted = sorted(HORARIOS, key=lambda x: -x['irr'])[:5]
hor_rows = ""
for h2 in sorted(HORARIOS, key=lambda x: x['h']):
    tot_irr = h2['atd'] + h2['adi']
    if tot_irr == 0 and h2['perd'] == 0: continue
    hor_rows += f"""<tr>
      <td>{h2['h']:02d}:00–{h2['h']:02d}:59</td>
      <td class="r">{fn(h2['v'])}</td>
      <td class="r" style="{'color:#DC2626;font-weight:700' if h2['perd']>10 else ''}">{h2['perd']}</td>
      <td class="r" style="{'color:#D97706;font-weight:700' if h2['atd']>30 else ''}">{h2['atd']}</td>
      <td class="r">{h2['adi']}</td>
      <td class="r"><strong style="color:{'#DC2626' if tot_irr>50 else '#D97706' if tot_irr>20 else '#374151'}">{tot_irr}</strong></td>
    </tr>"""

P_HORARIOS = f"""<div class="page">
  {hdr("Análise por Faixa Horária")}
  <div class="sec">Concentração de Irregularidades por Hora — PT (Atrasos + Adiantamentos)</div>
  {svg_heat_horas(HORARIOS)}
  <div style="margin-top:10px">
    <div class="sec">Detalhamento por Hora</div>
    <table>
      <thead><tr><th>Faixa</th><th class="r">Viagens</th><th class="r">Perdidas</th>
        <th class="r">Atrasos</th><th class="r">Adiant.</th><th class="r">Total Irr.</th></tr></thead>
      <tbody>{hor_rows}</tbody>
    </table>
  </div>
  {ftr(pg_num)}
</div>"""
pg_num += 1

# ── Página: Saída de Garagem + Recolhimento ──────────────────────────────────
sg = SAIDA_GAR
sg_pct_real = round(100*sg['real']/sg['tot'],1) if sg['tot'] else 0
sg_pct_pont = round(100*sg['pont']/sg['real'],1) if sg['real'] else 0
sg_pct_atrs = round(100*sg['atras']/sg['real'],1) if sg['real'] else 0
sg_pct_adt  = round(100*sg['adiant']/sg['real'],1) if sg['real'] else 0

sg_linhas_rows = ""
for sl in SAIDA_GAR_LINHAS:
    pct = round(100*sl['atras']/sl['tot'],1) if sl['tot'] else 0
    sg_linhas_rows += f"<tr><td>{sl['l']}</td><td class='r'>{sl['tot']}</td><td class='r' style='color:#DC2626;font-weight:700'>{sl['atras']}</td><td class='r'>{fp(pct)}</td></tr>"

rec_rows = ""
STATUS_PT = {"DENTRO_DO_PADRAO":"No Prazo","ACIMA_DO_PROGRAMADO":"Atrasado",
             "ABAIXO_DO_PROGRAMADO":"Adiantado","NAO_REALIZADA_COMPLETA":"Incompleto"}
for r2 in RECOLHE:
    pct = round(100*r2['n']/rec_tot,1) if rec_tot else 0
    rec_rows += f"<tr><td>{STATUS_PT.get(r2['status'],r2['status'])}</td><td class='r'>{r2['n']}</td><td class='r'>{fp(pct)}</td><td class='r'>{round(abs(r2['avg']),1)} min</td></tr>"

rec_linhas_rows = ""
for rl in RECOLHE_LINHAS:
    pct = round(100*rl['atras']/rl['n'],1) if rl['n'] else 0
    rec_linhas_rows += f"<tr><td>{rl['l']}</td><td class='r'>{rl['n']}</td><td class='r' style='color:#DC2626;font-weight:700'>{rl['atras']}</td><td class='r'>{fp(pct)}</td></tr>"

P_GARAGEM = f"""<div class="page">
  {hdr("Saída de Garagem e Recolhimento")}
  <div class="sec">Saída de Garagem</div>
  <div class="kpi-row">
    <div class="kpi"><div class="kpi-l">Programadas</div><div class="kpi-v sm" style="color:#1E40AF">{sg['tot']}</div></div>
    <div class="kpi"><div class="kpi-l">Realizadas</div><div class="kpi-v sm" style="color:#15803D">{sg['real']} <span style="font-size:11px">({sg_pct_real}%)</span></div></div>
    <div class="kpi"><div class="kpi-l">Pontuais</div><div class="kpi-v sm" style="color:#15803D">{sg['pont']} <span style="font-size:11px">({sg_pct_pont}%)</span></div></div>
    <div class="kpi"><div class="kpi-l">Atrasadas</div><div class="kpi-v sm" style="color:#DC2626">{sg['atras']} <span style="font-size:11px">({sg_pct_atrs}%)</span></div></div>
    <div class="kpi"><div class="kpi-l">Adiantadas</div><div class="kpi-v sm" style="color:#D97706">{sg['adiant']} <span style="font-size:11px">({sg_pct_adt}%)</span></div></div>
  </div>
  <div class="sec">Linhas com Mais Saídas Atrasadas</div>
  <table><thead><tr><th>Linha</th><th class="r">Total</th><th class="r">Atrasadas</th><th class="r">%</th></tr></thead>
    <tbody>{sg_linhas_rows or '<tr><td colspan="4" style="text-align:center;color:#6B7280">Sem dados</td></tr>'}</tbody></table>

  <div class="sec">Recolhimento de Garagem</div>
  <table><thead><tr><th>Status</th><th class="r">Qtd.</th><th class="r">%</th><th class="r">Diff. média</th></tr></thead>
    <tbody>{rec_rows or '<tr><td colspan="4" style="text-align:center;color:#6B7280">Sem dados</td></tr>'}</tbody></table>
  <div class="sec">Pior Linha de Recolhimento</div>
  <table><thead><tr><th>Linha</th><th class="r">Total</th><th class="r">Atrasados</th><th class="r">%</th></tr></thead>
    <tbody>{rec_linhas_rows or '<tr><td colspan="4" style="text-align:center;color:#6B7280">Sem dados</td></tr>'}</tbody></table>
  {ftr(pg_num)}
</div>"""
pg_num += 1

# ── Página: Aderência de Soltura ─────────────────────────────────────────────
_ADER_LABELS = {
    "Programado":    ("Programadas","#1E40AF"),
    "GNS":           ("GNS (Garagem Não Saiu)","#DC2626"),
    "Manutenção":    ("Manutenção","#D97706"),
    "Operação":      ("Oper. c/ restrição","#7C3AED"),
    "Faixa Amarela": ("Faixa Amarela","#B45309"),
}
if ADERENCIA:
    ader_kpis = ""
    for k, (lbl, cor) in _ADER_LABELS.items():
        v = ADERENCIA.get(k)
        if v is not None:
            vstr = f"{float(v):.1%}" if isinstance(v, float) else str(v)
            ader_kpis += f'<div class="kpi"><div class="kpi-l">{lbl}</div><div class="kpi-v sm" style="color:{cor}">{vstr}</div></div>'
    ader_body = f'<div class="kpi-row" style="flex-wrap:wrap">{ader_kpis}</div>'
    ader_body += f'<p style="font-size:10px;color:#6B7280;margin-top:6px">Fonte: planilha de aderência da rede (CCO) · {MES_NOME} {ANO}</p>'
else:
    ader_body = '<p style="font-size:11px;color:#D97706;padding:10px;background:#FFFBEB;border-radius:6px;border-left:3px solid #D97706">Planilha de aderência não disponível para este período. Verifique a rede CCO e rode novamente quando o arquivo estiver disponível.</p>'

P_ADERENCIA = f"""<div class="page">
  {hdr("Aderência de Soltura")}
  <div class="sec">Resultado Geral — {MES_NOME} {ANO}</div>
  {ader_body}
  {ftr(pg_num)}
</div>"""
pg_num += 1

# ── Página: Refeição ──────────────────────────────────────────────────────────
rf = REFEICAO
rf_pct = round(100*rf['com_int']/rf['tot'],1) if rf['tot'] else 0
P_REF = f"""<div class="page">
  {hdr("Refeição e Intervalo de Jornada")}
  <div class="sec">Intervalo Registrado — {MES_NOME} {ANO}</div>
  <div class="kpi-row">
    <div class="kpi"><div class="kpi-l">Registros totais</div><div class="kpi-v sm" style="color:#1E40AF">{fn(rf['tot'])}</div></div>
    <div class="kpi"><div class="kpi-l">Com intervalo</div><div class="kpi-v sm" style="color:#15803D">{fn(rf['com_int'])} <span style="font-size:11px">({rf_pct}%)</span></div></div>
    <div class="kpi"><div class="kpi-l">Duração média</div><div class="kpi-v sm" style="color:#1E40AF">{round(rf['avg_min'],0):.0f} min</div></div>
  </div>
  <p style="font-size:10px;color:#6B7280;margin-top:8px">
    Dados provenientes de <em>cco_indicadores_motorista</em> · cobertura de operadores que efetivamente
    trabalharam no período. Para detalhamento por terminal ou faixa, é necessário cruzamento com dados de escala.
  </p>
  {ftr(pg_num)}
</div>"""
pg_num += 1

# ── Página: Conclusão Executiva ───────────────────────────────────────────────
# Top positivos: terminais acima da meta
positivos = [f"<strong>{t}</strong> — CP {fp(td['cp'])} / PT {fp(td['pt'])}"
             for t, td in TERM_DATA.items() if td['cp'] >= META_CP and td['pt'] >= META_PT]
problemas = [f"<strong>{o['linha']}</strong> ({o['terminal']}) — {o['tipo']} {fp(o['pt'] if o['tipo']=='PT' else o['cp'])} com {o['oc']} ocorrências"
             for o in TOP_OPPS[:5] if o['pri'] in ['ALTA','MÉDIA']]
opps_text = [f"<strong>{o['linha']}</strong> — {o['tipo']} com {o['oc']} ocorrências (score {o['score']:.2f})"
             for o in TOP_OPPS[:5]]
atacar_text = [f"{o['ico']} <strong>{o['pri']}</strong> — {o['linha']} ({o['terminal']}, {o['tipo']})"
               for o in TOP_OPPS if o['pri']=='ALTA'][:5]

def ul(items, cor="#374151"):
    return "<ul style='padding-left:14px;margin:4px 0;font-size:11px;color:"+cor+"'>" + \
           "".join(f"<li style='margin-bottom:3px'>{i}</li>" for i in items) + "</ul>"

P_CONCLUSAO = f"""<div class="page">
  {hdr("Conclusão Executiva")}
  <div class="sec">Pontos Positivos</div>
  {ul(positivos or ["Verificar metas por terminal — nenhum terminal atingiu ambas as metas no período."], "#15803D")}

  <div class="sec">Principais Problemas</div>
  {ul(problemas or ["Nenhum problema de alta prioridade identificado."], "#DC2626")}

  <div class="sec">Principais Oportunidades</div>
  {ul(opps_text or ["Sem oportunidades identificadas."], "#D97706")}

  <div class="sec">Onde Atacar — Prioridades</div>
  {ul(atacar_text or ["✅ Sem prioridades ALTA no período."], "#1A3252")}

  <div style="margin-top:14px;padding:10px 14px;background:#F8FAFC;border-radius:7px;border:1px solid #E2E8F0">
    <div style="font-size:9px;font-weight:700;text-transform:uppercase;color:#6B7280;margin-bottom:4px">Metodologia de Priorização</div>
    <p style="font-size:10px;color:#374151;line-height:1.6">
      Score = Impacto (0,45) + Recorrência (0,40) + Concentração (0,15).
      🔴 ALTA ≥ 0,55 · 🟡 MÉDIA ≥ 0,28 · 🟢 BAIXA &lt; 0,28.
      Baseado em {fn(G['v'])} viagens analisadas no período {PERIODO_STR}.
    </p>
  </div>
  {ftr(pg_num)}
</div>"""

# ──────────────────────────────────────────────────────────────────────────────
# MONTA HTML FINAL
# ──────────────────────────────────────────────────────────────────────────────
HTML = f"""<!DOCTYPE html>
<html lang="pt-BR"><head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>Flash Report Mensal — {MES_NOME} {ANO}</title>
<style>{CSS}</style>
</head><body>
{P_CAPA}
{P_RESUMO}
{P_OPPS}
{''.join(PGS_TERM)}
{P_MOTIVOS}
{P_OFENSORES}
{P_HORARIOS}
{P_GARAGEM}
{P_ADERENCIA}
{P_REF}
{P_CONCLUSAO}
</body></html>"""

# ──────────────────────────────────────────────────────────────────────────────
# SAÍDA
# ──────────────────────────────────────────────────────────────────────────────
NOME = f"FLASH_REPORT_MENSAL_{MES:02d}_{ANO}"
OUT_HTML = os.path.join(r"C:\Users\monit\AppData\Local\Temp", f"{NOME}.html")
OUT_PDF  = os.path.join(SAIDA_DIR, f"{NOME}.pdf")

with open(OUT_HTML, "w", encoding="utf-8") as f:
    f.write(HTML)
print(f"HTML: {OUT_HTML} ({os.path.getsize(OUT_HTML)//1024} KB)")

if PREVIEW:
    import webbrowser
    webbrowser.open(OUT_HTML)
    print("[PREVIEW] Abrindo no navegador. PDF e envio ignorados.")
    sys.exit(0)

# PDF via Edge headless
_edge_candidates = [
    r"C:\Program Files (x86)\Microsoft\Edge\Application\msedge.exe",
    r"C:\Program Files\Microsoft\Edge\Application\msedge.exe",
]
_edge_candidates += glob.glob(r"C:\Program Files*\Microsoft\Edge\Application\msedge.exe")
_browser = next((c for c in _edge_candidates if os.path.exists(c)), None)
if _browser:
    _url = "file:///" + OUT_HTML.replace("\\", "/").replace(" ", "%20")
    result = subprocess.run(
        [_browser, "--headless=new", "--disable-gpu", "--no-sandbox",
         "--disable-extensions", f"--print-to-pdf={OUT_PDF}",
         "--print-to-pdf-no-header", _url],
        timeout=60, capture_output=True
    )
    if os.path.exists(OUT_PDF):
        print(f"PDF: {OUT_PDF} ({os.path.getsize(OUT_PDF)//1024} KB)")
    else:
        print(f"ERRO PDF: {result.stderr.decode(errors='ignore')[:200]}")
else:
    print("Edge não encontrado. PDF não gerado.")

if ENVIAR:
    # WhatsApp via Fonnte + gofile.io (mesma lógica do flash_diario.py)
    import requests
    FONNTE_TOKEN = _env("FONNTE_TOKEN") or ""
    WHATSAPP_TO  = [n.strip() for n in (_env("WHATSAPP_TO") or "").split(",") if n.strip()]

    if os.path.exists(OUT_PDF) and FONNTE_TOKEN and WHATSAPP_TO:
        # Upload PDF para gofile.io
        try:
            srv = requests.get("https://api.gofile.io/servers", timeout=10).json()
            server = srv["data"]["servers"][0]["name"]
            with open(OUT_PDF, "rb") as f2:
                up = requests.post(f"https://{server}.gofile.io/contents/uploadfile",
                                   files={"file": (os.path.basename(OUT_PDF), f2)}, timeout=60).json()
            link_pdf = up["data"]["downloadPage"]
        except Exception as e:
            link_pdf = f"PDF: {OUT_PDF}"
            print(f"gofile.io: {e}")

        msg = (f"📊 *FLASH REPORT MENSAL — {MES_NOME.upper()} {ANO}*\n"
               f"Período: {INICIO.strftime('%d/%m')} a {FIM.strftime('%d/%m/%Y')}\n\n"
               f"🏁 *CP:* {fp(G['cp'])} (meta {META_CP}%)\n"
               f"⏱ *PT:* {fp(G['pt'])} (meta {META_PT}%)\n\n"
               f"📌 Previstas: {fn(G['v'])} | Perdidas: {G['p']} | Atrasos: {G['atd']}\n\n"
               f"📄 {link_pdf}\n\n"
               f"_Atualizado em {GERADO_EM}_")

        for numero in WHATSAPP_TO:
            numero = numero.strip()
            if not numero: continue
            try:
                r2 = requests.post("https://api.fonnte.com/send",
                    headers={"Authorization": FONNTE_TOKEN},
                    data={"target": numero, "message": msg}, timeout=15)
                print(f"WhatsApp {numero}: {r2.status_code}")
            except Exception as e:
                print(f"WhatsApp {numero}: ERRO {e}")
    else:
        print("Envio WhatsApp pulado (PDF ausente, token ou números não configurados).")

print("Concluído.")
